#!/usr/bin/env python3
"""
Generate P6 XLSX Import File from WBS Summary Data

This script populates a P6 import template with data from a source file,
while strictly preserving the template's structure and additional tabs.

Features:
- Copies the P6 import template to preserve all sheets and formatting
- Maps data from WBS Summary to TASK sheet
- Preserves Row 1 (DB Keys) and Row 2 (User Headers) during population
- Deletes Row 2 (User Header) in final output
- Auto-generates Activity Codes (A1000, A1010, A1020, etc.)
- Ensures all data is formatted as text to prevent Excel auto-formatting
- Date-stamped output filename

Author: Senior Python Developer & P6 Data Engineer
Date: 2026-01-20
"""

import sys
from pathlib import Path
from datetime import datetime
import shutil
import logging
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s | %(levelname)-8s | %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger('p6_xlsx_generator')


class P6XLSXGenerator:
    """
    Generates P6 XLSX import file from WBS Summary data.
    """

    def __init__(self, template_path: Path, source_path: Path, output_dir: Path, project_name: str):
        """
        Initialize the generator.

        Args:
            template_path: Path to P6 import template XLSX
            source_path: Path to WBS Summary source data XLSX
            output_dir: Directory for output file
            project_name: Project name for output filename
        """
        self.template_path = template_path
        self.source_path = source_path
        self.output_dir = output_dir
        self.project_name = project_name
        
        # Generate output filename with date stamp
        date_stamp = datetime.now().strftime("%Y-%m-%d")
        self.output_filename = f"{project_name}_Import_{date_stamp}.xlsx"
        self.output_path = output_dir / self.output_filename

    def run(self):
        """Main execution flow."""
        logger.info("=" * 80)
        logger.info("P6 XLSX Import Generator")
        logger.info("=" * 80)

        try:
            # Step 1: Load source data
            source_df = self._load_source_data()

            # Step 2: Copy template to output
            self._copy_template()

            # Step 3: Populate TASK sheet
            self._populate_task_sheet(source_df)

            # Step 4: Delete Row 2 (User Header reference)
            self._delete_reference_row()

            logger.info("=" * 80)
            logger.info(f"✓ Successfully generated P6 import file: {self.output_path}")
            logger.info("=" * 80)

        except Exception as e:
            logger.error(f"Failed to generate P6 import file: {e}")
            raise

    def _load_source_data(self) -> pd.DataFrame:
        """Load and prepare source data from WBS Summary Excel."""
        logger.info(f"Loading source data from: {self.source_path}")

        df = pd.read_excel(self.source_path)
        logger.info(f"Loaded {len(df)} rows from source")
        logger.debug(f"Columns: {df.columns.tolist()}")

        # Filter out rows with empty WBS names
        original_count = len(df)
        df = df[
            df['WBS'].notna() & 
            (df['WBS'].astype(str).str.strip() != '')
        ]
        filtered_count = len(df)
        
        if original_count - filtered_count > 0:
            logger.info(f"Filtered out {original_count - filtered_count} rows with empty WBS names")
        logger.info(f"Total rows to process: {filtered_count}")

        # Generate automatic Activity Codes (A1000, A1010, A1020, etc.)
        activity_codes = [f"A{1000 + (i * 10)}" for i in range(len(df))]
        df['Activity Code'] = activity_codes
        logger.info(f"Generated {len(activity_codes)} activity codes (A1000 to A{1000 + (len(df)-1) * 10})")

        # Parse dates
        df['Start'] = pd.to_datetime(df['Start'], errors='coerce')
        df['Finish'] = pd.to_datetime(df['Finish'], errors='coerce')

        # Log date parsing issues
        invalid_start = df['Start'].isna().sum()
        invalid_finish = df['Finish'].isna().sum()
        if invalid_start > 0:
            logger.warning(f"{invalid_start} rows have invalid Start dates")
        if invalid_finish > 0:
            logger.warning(f"{invalid_finish} rows have invalid Finish dates")

        return df

    def _copy_template(self):
        """Create a copy of the template file."""
        logger.info(f"Copying template to: {self.output_path}")
        
        # Ensure output directory exists
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        # Copy template
        shutil.copy2(self.template_path, self.output_path)
        logger.info("Template copied successfully")

    def _populate_task_sheet(self, df: pd.DataFrame):
        """
        Populate the TASK sheet with source data.
        
        CRITICAL: Preserves Row 1 (DB Keys) and Row 2 (User Headers).
        Data is inserted starting at Row 3.
        All data is formatted as text to prevent Excel auto-formatting.
        """
        logger.info("Populating TASK sheet with source data...")

        # Load the copied workbook
        wb = load_workbook(self.output_path)
        ws = wb['TASK']

        # Get column mapping from Row 1 (DB Keys)
        db_keys = {cell.value: cell.column for cell in ws[1] if cell.value}
        logger.debug(f"DB Key columns: {db_keys}")

        # Column mapping: Source column -> Template DB Key
        column_mapping = {
            'Activity Code': 'task_code',      # Auto-generated activity code
            'WBS': 'task_name',                 # Activity name
            'Remaining Duration': 'remain_drtn_hr_cnt',  # Remaining duration
            'Start': 'start_date',              # Start date
            'Finish': 'end_date',               # Finish date
        }

        # Status for all activities (Not Started by default)
        default_status = 'TK_NotStart'

        # Insert data starting at Row 3
        start_row = 3
        tasks_written = 0

        for idx, row in df.iterrows():
            current_row = start_row + tasks_written

            # Write each mapped field
            for source_col, db_key in column_mapping.items():
                if db_key in db_keys:
                    col_idx = db_keys[db_key]
                    value = row.get(source_col, '')
                    
                    # Format value as string to prevent auto-formatting
                    if pd.notna(value):
                        if isinstance(value, datetime):
                            # Format dates as string
                            value = value.strftime('%d-%b-%y')
                        else:
                            value = str(value)
                            # Strip leading spaces from activity names to remove indentation
                            if db_key == 'task_name':
                                value = value.lstrip()
                            # Strip 'd' suffix from remaining duration (e.g., '605.9d' -> '605.9')
                            if db_key == 'remain_drtn_hr_cnt':
                                value = value.rstrip('d').strip()
                    else:
                        value = ''
                    
                    ws.cell(row=current_row, column=col_idx, value=value)

            # Set status code (human-readable format)
            if 'status_code' in db_keys:
                # Determine status based on start date
                start_date = row.get('Start')
                finish_date = row.get('Finish')
                
                if pd.notna(finish_date) and finish_date <= datetime.now():
                    status = 'Complete'
                elif pd.notna(start_date) and start_date <= datetime.now():
                    status = 'In Progress'
                else:
                    status = 'Not Started'
                ws.cell(row=current_row, column=db_keys['status_code'], value=status)

            tasks_written += 1

        # Save workbook
        wb.save(self.output_path)
        logger.info(f"Wrote {tasks_written} tasks to TASK sheet")

        # Verify other sheets are preserved
        logger.info(f"Sheets preserved: {wb.sheetnames}")

    def _delete_reference_row(self):
        """Delete Row 2 (User Header reference row) from TASK sheet."""
        logger.info("Deleting Row 2 (User Header reference) from TASK sheet...")

        wb = load_workbook(self.output_path)
        ws = wb['TASK']

        # Delete row 2
        ws.delete_rows(2)

        wb.save(self.output_path)
        logger.info("Row 2 deleted successfully")


def main():
    """Main entry point."""
    # Define file paths
    script_dir = Path(__file__).parent
    base_dir = script_dir.parent

    # Input/Output paths
    template_path = base_dir / "input" / "P6_Import_template.xlsx"
    source_path = base_dir / "input" / "WBS Summary.xlsx"
    output_dir = base_dir / "output"
    
    # Project name for output filename
    project_name = "19282_Summary_Schedule"

    # Verify input files exist
    if not template_path.exists():
        logger.error(f"P6 Import Template not found: {template_path}")
        sys.exit(1)

    if not source_path.exists():
        logger.error(f"WBS Summary not found: {source_path}")
        sys.exit(1)

    # Run generator
    generator = P6XLSXGenerator(
        template_path=template_path,
        source_path=source_path,
        output_dir=output_dir,
        project_name=project_name
    )

    generator.run()


if __name__ == '__main__':
    main()
